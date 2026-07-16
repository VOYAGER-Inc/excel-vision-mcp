"""
Core Excel reading engine for the MCP Server.

Handles workbook parsing, sheet enumeration, cell data extraction with chunking
support for large files, and merged cell detection. Designed to work with
openpyxl in read-only mode for memory efficiency on large workbooks.
"""

import logging
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from .image_extractor import count_images_in_archive

logger = logging.getLogger(__name__)


@dataclass
class SheetInfo:
    """
    Metadata about a single worksheet.

    Provides dimensional and structural information to help agents
    understand the sheet's layout before requesting full data.
    """

    name: str
    row_count: int
    col_count: int
    max_row: int
    max_col: int
    merged_cell_count: int = 0
    has_images: bool = False
    data_range: str = ""


@dataclass
class WorkbookSummary:
    """
    High-level overview of an entire workbook.

    Enables agents to quickly assess file contents, size, and structure
    before making targeted data requests.
    """

    file_path: str
    file_name: str
    file_size_bytes: int
    sheet_count: int
    sheets: list[SheetInfo] = field(default_factory=list)
    total_image_count: int = 0


@dataclass
class CellData:
    """
    Structured representation of a single cell's contents.

    Captures both value and positional information for accurate reporting.
    """

    coordinate: str
    value: str | None
    row: int
    col: int
    is_merged: bool = False


@dataclass
class SheetData:
    """
    Paginated chunk of sheet data.

    Supports pagination metadata so agents can request subsequent chunks
    of large sheets without loading everything at once.
    """

    sheet_name: str
    cells: list[CellData] = field(default_factory=list)
    start_row: int = 1
    end_row: int = 0
    total_rows: int = 0
    total_cols: int = 0
    has_more: bool = False
    merged_cells: list[str] = field(default_factory=list)


def get_workbook_summary(file_path: str) -> WorkbookSummary:
    """
    Generate a comprehensive overview of the workbook structure.

    Opens the workbook in read-only mode to minimize memory usage,
    gathering sheet metadata, dimensions, and image counts.

    @param file_path: Absolute path to the .xlsx file.
    @return: WorkbookSummary with all sheet metadata.
    @throws FileNotFoundError: If file_path doesn't exist.
    @throws openpyxl.utils.exceptions.InvalidFileException: If not a valid xlsx.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    if not path.suffix.lower() in (".xlsx", ".xlsm"):
        raise ValueError(f"Unsupported file format: {path.suffix}. Only .xlsx/.xlsm supported.")

    file_size = path.stat().st_size
    total_images = count_images_in_archive(file_path)

    # Normal mode (not read_only) to get accurate max_row/max_column
    # read_only mode returns None for dimensions in some xlsx files
    wb = openpyxl.load_workbook(file_path, data_only=True)

    sheets: list[SheetInfo] = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        data_range = ""
        if max_row > 0 and max_col > 0:
            end_col_letter = get_column_letter(max_col)
            data_range = f"A1:{end_col_letter}{max_row}"

        merged_count = len(ws.merged_cells.ranges)

        sheets.append(SheetInfo(
            name=sname,
            row_count=max_row,
            col_count=max_col,
            max_row=max_row,
            max_col=max_col,
            merged_cell_count=merged_count,
            has_images=len(getattr(ws, "_images", [])) > 0,
            data_range=data_range,
        ))

    wb.close()

    summary = WorkbookSummary(
        file_path=file_path,
        file_name=path.name,
        file_size_bytes=file_size,
        sheet_count=len(sheets),
        sheets=sheets,
        total_image_count=total_images,
    )

    return summary


def _extract_sheet_data(
    ws,
    start_row: int = 1,
    max_rows: int = 200,
    start_col: int = 1,
    max_cols: int | None = None,
) -> SheetData:
    """
    Extract paginated cell data from an already-loaded worksheet.

    Shared by read_sheet_data and read_all_sheets_data so the workbook
    is loaded exactly once per file regardless of how many sheets are read.

    @param ws: Worksheet loaded in normal mode (accurate dimensions and merged ranges).
    @param start_row: 1-indexed starting row.
    @param max_rows: Maximum number of rows to return per chunk.
    @param start_col: 1-indexed starting column.
    @param max_cols: Maximum columns, or None for all.
    @return: SheetData with paginated cell contents.
    """
    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0

    merged_ranges: list[str] = []
    merged_cells_set: set[str] = set()
    for merged_range in ws.merged_cells.ranges:
        merged_ranges.append(str(merged_range))
        for row, col in merged_range.cells:
            merged_cells_set.add(f"{get_column_letter(col)}{row}")

    end_row = min(start_row + max_rows - 1, total_rows) if total_rows > 0 else start_row + max_rows - 1
    effective_max_cols = max_cols or total_cols

    cells: list[CellData] = []
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=min(start_col + effective_max_cols - 1, total_cols) if total_cols > 0 else None,
    ):
        for cell in row:
            if cell.value is not None:
                coord = f"{get_column_letter(cell.column)}{cell.row}"
                cells.append(CellData(
                    coordinate=coord,
                    value=str(cell.value),
                    row=cell.row,
                    col=cell.column,
                    is_merged=coord in merged_cells_set,
                ))

    return SheetData(
        sheet_name=ws.title,
        cells=cells,
        start_row=start_row,
        end_row=end_row,
        total_rows=total_rows,
        total_cols=total_cols,
        has_more=end_row < total_rows,
        merged_cells=merged_ranges,
    )


def read_sheet_data(
    file_path: str,
    sheet_name: str | None = None,
    start_row: int = 1,
    max_rows: int = 200,
    start_col: int = 1,
    max_cols: int | None = None,
) -> SheetData:
    """
    Read cell data from a sheet with pagination support.

    Loads the workbook once in normal mode (read_only mode returns None
    dimensions for some xlsx files) and paginates large sheets to prevent
    context overflow.

    @param file_path: Absolute path to the .xlsx file.
    @param sheet_name: Sheet to read, or None for the active/first sheet.
    @param start_row: 1-indexed starting row.
    @param max_rows: Maximum number of rows to return per chunk.
    @param start_col: 1-indexed starting column.
    @param max_cols: Maximum columns, or None for all.
    @return: SheetData with paginated cell contents.
    @throws FileNotFoundError: If file_path doesn't exist.
    @throws ValueError: If sheet_name doesn't exist in the workbook.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        ws_name = sheet_name or wb.sheetnames[0]
        if ws_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{ws_name}' not found. Available: {wb.sheetnames}")

        return _extract_sheet_data(
            wb[ws_name],
            start_row=start_row,
            max_rows=max_rows,
            start_col=start_col,
            max_cols=max_cols,
        )
    finally:
        wb.close()


def read_all_sheets_data(
    file_path: str,
    max_rows_per_sheet: int = 500,
) -> list[SheetData]:
    """
    Read data from all sheets in a workbook.

    Loads the workbook once and iterates through every sheet, applying
    per-sheet row limits to prevent memory exhaustion on large workbooks.

    @param file_path: Absolute path to the .xlsx file.
    @param max_rows_per_sheet: Maximum rows to read per sheet.
    @return: List of SheetData, one per sheet.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        return [
            _extract_sheet_data(wb[sname], max_rows=max_rows_per_sheet)
            for sname in wb.sheetnames
        ]
    finally:
        wb.close()


def search_in_workbook(
    file_path: str,
    query: str,
    sheet_name: str | None = None,
    case_sensitive: bool = False,
) -> list[dict]:
    """
    Search for text across all cells in the workbook.

    Performs a substring search across all (or specified) sheets, returning
    matching cells with surrounding context rows for agent analysis.

    @param file_path: Absolute path to the .xlsx file.
    @param query: Search string.
    @param sheet_name: Limit search to specific sheet, or None for all.
    @param case_sensitive: Whether search is case-sensitive.
    @return: List of match dicts with sheet, cell, value, and context.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets_to_search = [sheet_name] if sheet_name else wb.sheetnames
    search_query = query if case_sensitive else query.lower()

    matches: list[dict] = []
    max_matches = 100

    for sname in sheets_to_search:
        if sname not in wb.sheetnames:
            continue
        if len(matches) >= max_matches:
            break

        ws = wb[sname]
        for row in ws.iter_rows():
            if len(matches) >= max_matches:
                break

            for cell in row:
                if cell.value is not None:
                    cell_str = str(cell.value)
                    compare_str = cell_str if case_sensitive else cell_str.lower()

                    if search_query in compare_str:
                        coord = f"{get_column_letter(cell.column)}{cell.row}"
                        matches.append({
                            "sheet": sname,
                            "cell": coord,
                            "row": cell.row,
                            "col": cell.column,
                            "value": cell_str,
                        })

    wb.close()
    return matches
