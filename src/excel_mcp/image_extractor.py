"""
Dual-strategy image extractor for Excel files.

Uses two complementary methods to maximize image recovery:
1. openpyxl-image-loader: Maps images to specific cell positions
2. zipfile extraction: Catches all images from xl/media/ as fallback

The combined approach ensures no embedded images are missed while preserving
cell position metadata whenever possible.
"""

import hashlib
import logging
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

from .utils import (
    DEFAULT_MAX_HEIGHT,
    DEFAULT_MAX_WIDTH,
    DEFAULT_JPEG_QUALITY,
    ImageInfo,
    bytes_to_base64,
    image_to_base64,
)

logger = logging.getLogger(__name__)

# Supported image extensions inside xlsx archives
SUPPORTED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".emf", ".wmf"}


@dataclass
class ExtractionResult:
    """
    Aggregated result from dual-strategy image extraction.

    Separates cell-mapped images (with known positions) from orphan images
    (found in archive but not mapped to cells) for comprehensive reporting.
    """

    cell_mapped_images: list[ImageInfo] = field(default_factory=list)
    orphan_images: list[ImageInfo] = field(default_factory=list)
    total_count: int = 0
    errors: list[str] = field(default_factory=list)


def _extract_via_image_loader(
    file_path: str,
    sheet_name: str | None = None,
    max_width: int = DEFAULT_MAX_WIDTH,
    max_height: int = DEFAULT_MAX_HEIGHT,
    quality: int = DEFAULT_JPEG_QUALITY,
) -> list[ImageInfo]:
    """
    Extract images using openpyxl-image-loader with cell position mapping.

    This method provides accurate cell-to-image mapping but may miss images
    that aren't anchored to the standard drawing layer.

    @param file_path: Absolute path to the .xlsx file.
    @param sheet_name: Specific sheet name, or None for all sheets.
    @param max_width: Maximum image width after optimization.
    @param max_height: Maximum image height after optimization.
    @param quality: JPEG compression quality.
    @return: List of ImageInfo with cell references populated.
    """
    import openpyxl
    from openpyxl_image_loader import SheetImageLoader

    results: list[ImageInfo] = []
    wb = openpyxl.load_workbook(file_path)

    sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames

    image_index = 0
    for sname in sheets_to_process:
        if sname not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping", sname)
            continue

        sheet = wb[sname]

        try:
            loader = SheetImageLoader(sheet)
        except Exception as e:
            logger.debug("No images in sheet '%s': %s", sname, e)
            continue

        # The loader keys its internal map by anchor cell, so iterating it is
        # O(images). Fall back to scanning the used range if the private
        # attribute ever disappears in a future release.
        anchor_map = getattr(loader, "_images", None)
        if anchor_map is not None:
            anchor_cells = list(anchor_map.keys())
        else:
            anchor_cells = [
                cell.coordinate
                for row in sheet.iter_rows()
                for cell in row
                if loader.image_in(cell.coordinate)
            ]

        for coord in anchor_cells:
            try:
                pil_image = loader.get(coord)
                original_w, original_h = pil_image.size

                encoded, mime_type, proc_w, proc_h = image_to_base64(
                    pil_image, max_width, max_height, quality
                )

                results.append(ImageInfo(
                    data_base64=encoded,
                    mime_type=mime_type,
                    cell_reference=coord,
                    sheet_name=sname,
                    original_width=original_w,
                    original_height=original_h,
                    processed_width=proc_w,
                    processed_height=proc_h,
                    index=image_index,
                ))
                image_index += 1
                logger.debug(
                    "Extracted image at %s!%s (%dx%d)", sname, coord, original_w, original_h
                )
            except Exception as e:
                logger.debug("Error extracting image at %s!%s: %s", sname, coord, e)

    wb.close()
    return results


def _extract_via_zipfile(
    file_path: str,
    max_width: int = DEFAULT_MAX_WIDTH,
    max_height: int = DEFAULT_MAX_HEIGHT,
    quality: int = DEFAULT_JPEG_QUALITY,
) -> list[ImageInfo]:
    """
    Extract all images from the xlsx ZIP archive's xl/media/ directory.

    This fallback method catches every embedded image but cannot determine
    which cell an image belongs to. Used to find images missed by the
    cell-mapping approach.

    @param file_path: Absolute path to the .xlsx file.
    @param max_width: Maximum image width after optimization.
    @param max_height: Maximum image height after optimization.
    @param quality: JPEG compression quality.
    @return: List of ImageInfo with original_filename populated (no cell refs).
    """
    results: list[ImageInfo] = []

    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            media_files = [
                name for name in zf.namelist()
                if name.startswith("xl/media/")
                and Path(name).suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
            ]

            for idx, media_path in enumerate(sorted(media_files)):
                try:
                    raw_bytes = zf.read(media_path)
                    filename = Path(media_path).name

                    encoded, mime_type, orig_w, orig_h, proc_w, proc_h = bytes_to_base64(
                        raw_bytes, max_width, max_height, quality
                    )

                    results.append(ImageInfo(
                        data_base64=encoded,
                        mime_type=mime_type,
                        original_filename=filename,
                        original_width=orig_w,
                        original_height=orig_h,
                        processed_width=proc_w,
                        processed_height=proc_h,
                        file_size_bytes=len(raw_bytes),
                        index=idx,
                    ))
                    logger.debug("Extracted archive image: %s (%dx%d)", filename, orig_w, orig_h)

                except Exception as e:
                    logger.warning("Failed to process %s: %s", media_path, e)

    except zipfile.BadZipFile:
        logger.error("File is not a valid xlsx/zip archive: %s", file_path)
    except Exception as e:
        logger.error("Zipfile extraction error: %s", e)

    return results


def extract_all_images(
    file_path: str,
    sheet_name: str | None = None,
    max_width: int = DEFAULT_MAX_WIDTH,
    max_height: int = DEFAULT_MAX_HEIGHT,
    quality: int = DEFAULT_JPEG_QUALITY,
) -> ExtractionResult:
    """
    Extract images using dual strategy: cell-mapping first, then archive fallback.

    Merges results from both methods, deduplicating by image content hash.
    Cell-mapped images take priority; archive-only images are marked as orphans.

    @param file_path: Absolute path to the .xlsx file.
    @param sheet_name: Specific sheet name, or None for all sheets.
    @param max_width: Maximum image width after optimization.
    @param max_height: Maximum image height after optimization.
    @param quality: JPEG compression quality.
    @return: ExtractionResult with cell-mapped and orphan images.
    """
    result = ExtractionResult()

    # Strategy 1: Cell-mapped extraction
    try:
        cell_images = _extract_via_image_loader(
            file_path, sheet_name, max_width, max_height, quality
        )
        result.cell_mapped_images = cell_images
        logger.info("Cell-mapper found %d images", len(cell_images))
    except Exception as e:
        error_msg = f"Cell-mapping extraction failed: {e}"
        result.errors.append(error_msg)
        logger.warning(error_msg)

    # Strategy 2: Archive extraction (catches missed images)
    try:
        archive_images = _extract_via_zipfile(
            file_path, max_width, max_height, quality
        )
        logger.info("Archive extraction found %d images", len(archive_images))

        # Deduplicate: keep archive images not already found via cell-mapping.
        # Hash the full content — image headers (e.g. JPEG quantization tables)
        # are often identical across distinct images, so a prefix isn't enough.
        cell_hashes = {
            hashlib.sha256(img.data_base64.encode("ascii")).digest()
            for img in result.cell_mapped_images
        }

        for img in archive_images:
            if hashlib.sha256(img.data_base64.encode("ascii")).digest() not in cell_hashes:
                result.orphan_images.append(img)

        if result.orphan_images:
            logger.info(
                "Found %d orphan images not mapped to cells", len(result.orphan_images)
            )

    except Exception as e:
        error_msg = f"Archive extraction failed: {e}"
        result.errors.append(error_msg)
        logger.warning(error_msg)

    result.total_count = len(result.cell_mapped_images) + len(result.orphan_images)
    return result


def count_images_in_archive(file_path: str) -> int:
    """
    Quickly count images in the xlsx archive without extracting them.

    @param file_path: Absolute path to the .xlsx file.
    @return: Number of image files found in xl/media/.
    """
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            return len([
                name for name in zf.namelist()
                if name.startswith("xl/media/")
                and Path(name).suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
            ])
    except Exception:
        return 0
