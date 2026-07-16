"""
Shared pytest fixtures for the Excel MCP Server test suite.

All test workbooks are generated on the fly in pytest's tmp_path, so the
suite is fully self-contained: no real-world files, no machine-specific
paths, and nothing sensitive to leak into the repository.
"""

import openpyxl
import pytest
from openpyxl.drawing.image import Image as XlsxImage
from PIL import Image as PILImage


def _make_png(path, size, color):
    """Create a solid-color PNG file for embedding into test workbooks."""
    PILImage.new("RGB", size, color).save(path, format="PNG")
    return path


@pytest.fixture
def sample_workbook(tmp_path):
    """
    Workbook with two sheets, typical cell data, and a merged range.

    Sheet "Data": 3-column header + 30 data rows, A1:C1 merged title row.
    Sheet "Notes": a few sparse cells.
    """
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Quarterly Report"
    ws.merge_cells("A1:C1")
    ws.append(["ID", "Name", "Amount"])
    for i in range(1, 31):
        ws.append([i, f"Item {i}", i * 10])

    notes = wb.create_sheet("Notes")
    notes["B2"] = "Reviewed by QA"
    notes["B3"] = "Approved"

    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def image_workbook(tmp_path):
    """
    Workbook with two distinct images anchored at Diagrams!B2 and Diagrams!D5.

    The images differ in color and size so their processed content differs,
    which exercises the dedup logic between cell-mapped and archive results.
    """
    path = tmp_path / "images.xlsx"
    red_png = _make_png(tmp_path / "red.png", (64, 48), (255, 0, 0))
    blue_png = _make_png(tmp_path / "blue.png", (100, 80), (0, 0, 255))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diagrams"
    ws["A1"] = "Diagram sheet"
    ws.add_image(XlsxImage(str(red_png)), "B2")
    ws.add_image(XlsxImage(str(blue_png)), "D5")
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def large_image_workbook(tmp_path):
    """Workbook with one image larger than the default optimization bounds."""
    path = tmp_path / "large_image.xlsx"
    big_png = _make_png(tmp_path / "big.png", (1600, 800), (0, 128, 0))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.add_image(XlsxImage(str(big_png)), "A1")
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def not_an_xlsx(tmp_path):
    """A file with .xlsx extension that is not a valid ZIP archive."""
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"this is not a zip archive")
    return path
